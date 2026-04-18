"""Generate Delhi_PS_CRM_Quotation.xlsx with formula-driven totals.

Every Total Cost cell uses =Quantity*UnitCost.
Phase subtotals use =SUM(range).
Section 2 phase costs reference Section 3 subtotal cells.
Unit Economics derive from subtotal cells — zero hardcoded results.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_FILENAME = "Delhi_PS_CRM_Quotation.xlsx"

# ─── Style constants ───────────────────────────────────────────────
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1F3864")
FONT_SUBTITLE = Font(name="Calibri", size=12, color="1F3864")
FONT_META = Font(name="Calibri", size=11, color="1F3864")
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
FONT_TH = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_BOLD_NAVY = Font(name="Calibri", size=11, bold=True, color="1F3864")
FONT_BOLD_NAVY_12 = Font(name="Calibri", size=12, bold=True, color="1F3864")

FILL_SECTION = PatternFill("solid", fgColor="1F3864")
FILL_TH = PatternFill("solid", fgColor="2E4F8A")
FILL_SUBTOTAL = PatternFill("solid", fgColor="BDD7EE")
FILL_ALT = PatternFill("solid", fgColor="EBF3FF")

AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
AR = Alignment(horizontal="right", vertical="top", wrap_text=True)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
ACL = Alignment(horizontal="left", vertical="center", wrap_text=True)

CURRENCY_FMT = "#,##,##0"

THIN = Side(style="thin", color="A6A6A6")
THICK = Side(style="thick", color="1F3864")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Helpers ───────────────────────────────────────────────────────

def _col_widths(ws, lo: int = 12, hi: int = 70) -> None:
    """Auto-size columns between lo and hi character widths."""
    widths: dict[int, int] = {}
    for r in ws.iter_rows():
        for c in r:
            if c.value is None:
                continue
            text = str(c.value)
            ml = max((len(ln) for ln in text.splitlines()), default=len(text))
            widths[c.column] = max(widths.get(c.column, 0), ml)
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = max(lo, min(hi, w + 2))


def _border_row(ws, r: int, c1: int, c2: int) -> None:
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = BORDER


def _fill_row(ws, r: int, c1: int, c2: int, fill: PatternFill) -> None:
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).fill = fill


def _section_hdr(ws, r: int, c1: int, c2: int, title: str) -> None:
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=title)
    cell.font = FONT_SECTION
    cell.alignment = ACL
    _fill_row(ws, r, c1, c2, FILL_SECTION)


def _table_hdr(ws, r: int, c1: int, c2: int) -> None:
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_TH
        cell.font = FONT_TH
        cell.alignment = AC
        cell.border = BORDER


def _thick_box(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border or Border()
            cell.border = Border(
                left=THICK if c == c1 else b.left,
                right=THICK if c == c2 else b.right,
                top=THICK if r == r1 else b.top,
                bottom=THICK if r == r2 else b.bottom,
            )


def _cur(cell, value) -> None:
    """Set cell with Indian currency format and right alignment."""
    cell.value = value
    cell.number_format = CURRENCY_FMT
    cell.alignment = AR


# ─── Build ─────────────────────────────────────────────────────────

def build_workbook() -> Workbook:  # noqa: C901 — single builder is intentional
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.freeze_panes = "A4"

    # ================================================================
    # ROWS 1-3 — Title block
    # ================================================================
    ws.merge_cells("A1:H1")
    ws["A1"] = "Delhi PS-CRM — Project Cost Quotation"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = AC

    ws.merge_cells("A2:H2")
    ws["A2"] = "AI-Powered WhatsApp Civic Complaint Management System"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = AC

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        "Team: Delhi PS-CRM | Document: Cost Quotation v2.0 "
        "| Excluding all applicable taxes"
    )
    ws["A3"].font = FONT_META
    ws["A3"].alignment = AC

    row = 5

    # ================================================================
    # SECTION 1 — PROJECT OVERVIEW HEADER
    # ================================================================
    s1 = row
    _section_hdr(ws, row, 1, 8, "SECTION 1 — PROJECT OVERVIEW HEADER")
    row += 1

    info_items = [
        ("Target Scale", "1,00,000 complaints/month across 272 Delhi wards"),
        (
            "Technology Stack",
            "FastAPI, Gemini 2.5 Flash-Lite, Gradient Boosting ML, "
            "Azure Cloud, WhatsApp Business API, SendGrid",
        ),
        ("Deployment Timeline", "2 months"),
        ("Languages Supported", "Hindi and English"),
    ]
    for k, v in info_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"{k}:").font = FONT_BOLD_NAVY
        ws.cell(row=row, column=1).alignment = AL
        ws.cell(row=row, column=3, value=v).alignment = AL
        _border_row(ws, row, 1, 8)
        row += 1

    _thick_box(ws, s1, row - 1, 1, 8)
    row += 2

    # ================================================================
    # SECTION 2 — PHASE-WISE COST SUMMARY
    # (Phase Cost column filled via formula after Section 3 is built)
    # ================================================================
    s2 = row
    _section_hdr(ws, row, 1, 6, "SECTION 2 — PHASE-WISE COST SUMMARY")
    row += 1

    for i, h in enumerate(
        [
            "Phase",
            "Duration",
            "Description",
            "Key Activities",
            "Est. Complaints/Month",
            "Phase Cost (Rs.)",
        ],
        start=1,
    ):
        ws.cell(row=row, column=i, value=h)
    _table_hdr(ws, row, 1, 6)
    row += 1

    summaries = [
        (
            "Phase 1", "Weeks 1-2", "Development and Testing",
            "Backend complete, WhatsApp bot on test numbers, ML model "
            "integrated, staging environment, internal QA",
            "0",
        ),
        (
            "Phase 2", "Weeks 3-4", "Pilot Deployment",
            "Live in 5 Delhi wards, real citizen complaints, bug fixes, "
            "officer onboarding",
            "1,000",
        ),
        (
            "Phase 3", "Weeks 5-6", "Partial Rollout",
            "Expand to 50 wards, department routing live, ML escalation "
            "running on real data",
            "20,000",
        ),
        (
            "Phase 4", "Weeks 7-8", "Full Delhi Deployment",
            "All 272 wards live, 1,00,000 complaints/month, full system "
            "operational",
            "1,00,000",
        ),
        (
            "Phase 5", "Month 3 onwards", "Steady State Operations",
            "Handed to Delhi government IT team. Maintenance only, "
            "quarterly ML retraining",
            "1,00,000",
        ),
    ]

    sum_rows: dict[str, int] = {}  # phase name -> summary-table row
    sum_start = row
    for ph, dur, desc, acts, est in summaries:
        ws.cell(row=row, column=1, value=ph)
        ws.cell(row=row, column=2, value=dur)
        ws.cell(row=row, column=3, value=desc)
        ws.cell(row=row, column=4, value=acts)
        ws.cell(row=row, column=5, value=est).alignment = AR
        # Column 6 placeholder — filled with formula after Section 3
        sum_rows[ph] = row
        row += 1
    sum_end = row - 1

    # Zebra + borders for summary rows
    for r in range(sum_start, sum_end + 1):
        _border_row(ws, r, 1, 6)
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = AL
        ws.cell(row=r, column=5).alignment = AR
        if (r - sum_start) % 2 == 1:
            _fill_row(ws, r, 1, 6, FILL_ALT)

    # Grand Total row (Phases 1-4)
    gt_row = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="Total Project Cost Phases 1-4").font = FONT_BOLD_NAVY
    ws.cell(row=row, column=1).alignment = AL
    _border_row(ws, row, 1, 6)
    row += 1

    # Annual Maintenance row
    am_row = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="Annual Maintenance from Month 3").font = FONT_BOLD_NAVY
    ws.cell(row=row, column=1).alignment = AL
    _border_row(ws, row, 1, 6)
    s2e = row
    _thick_box(ws, s2, s2e, 1, 6)
    row += 2

    # ================================================================
    # SECTION 3 — DETAILED LINE-ITEM BREAKDOWN
    # ================================================================
    s3 = row
    _section_hdr(ws, row, 1, 8, "SECTION 3 — DETAILED LINE-ITEM BREAKDOWN")
    row += 1

    for i, h in enumerate(
        [
            "Phase",
            "Category",
            "Item",
            "Specifications and Details",
            "Quantity / Duration",
            "Unit Cost (Rs.)",
            "Total Cost (Rs.)",
            "Notes",
        ],
        start=1,
    ):
        ws.cell(row=row, column=i, value=h)
    _table_hdr(ws, row, 1, 8)
    row += 1

    detail_start = row
    sub_rows: dict[str, int] = {}   # "Phase N" -> subtotal row in Section 3
    first_rows: dict[str, int] = {} # "Phase N" -> first item row
    item_row_nums: list[int] = []   # all item (non-subtotal) row numbers

    # ── write a single line item ──
    def item(
        phase: str,
        cat: str,
        name: str,
        specs: str,
        qty,       # int, float, or str formula like "=5*0.5"
        unit,      # int or float — the raw unit cost number
        notes: str = "",
    ) -> None:
        nonlocal row
        if phase not in first_rows:
            first_rows[phase] = row

        ws.cell(row=row, column=1, value=phase).alignment = AL
        ws.cell(row=row, column=2, value=cat).alignment = AL
        ws.cell(row=row, column=3, value=name).alignment = AL
        ws.cell(row=row, column=4, value=specs).alignment = AL

        # Quantity cell (col E) — raw hardcoded input
        # Always use '0.##' so 1→"1", 0.5→"0.5", 2.5→"2.5" (no trailing dot)
        qc = ws.cell(row=row, column=5, value=qty)
        qc.number_format = "0.##"
        qc.alignment = AR

        # Unit Cost cell (col F) — raw hardcoded input
        uc = ws.cell(row=row, column=6, value=unit)
        if isinstance(unit, float) and unit < 1:
            uc.number_format = "0.00#"
        elif isinstance(unit, float):
            uc.number_format = "#,##0.00"
        else:
            uc.number_format = CURRENCY_FMT
        uc.alignment = AR

        # Total Cost cell (col G) — ALWAYS a formula
        _cur(ws.cell(row=row, column=7), f"=E{row}*F{row}")

        ws.cell(row=row, column=8, value=notes).alignment = AL

        item_row_nums.append(row)
        row += 1

    # ── write a phase subtotal ──
    def phase_subtotal(phase: str) -> None:
        nonlocal row
        fr = first_rows[phase]
        lr = row - 1
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = FILL_SUBTOTAL
            ws.cell(row=row, column=c).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"{phase} Subtotal").font = FONT_BOLD
        ws.cell(row=row, column=1).alignment = AL
        _cur(ws.cell(row=row, column=7), f"=SUM(G{fr}:G{lr})")
        ws.cell(row=row, column=7).font = FONT_BOLD
        sub_rows[phase] = row
        row += 1

    # ────────────────────────────────────────────────────────────
    # PHASE 1 — Development and Testing (Weeks 1-2)
    # ────────────────────────────────────────────────────────────
    item(
        "Phase 1", "Team", "Student Developer Stipends",
        "5 team members: backend, ML integration, architecture, testing, DevOps",
        "=5*0.5", 25000,
        "Half-month stipend weeks 1-2",
    )
    item(
        "Phase 1", "Infrastructure", "Azure Container Apps Staging",
        "Staging FastAPI backend for internal testing",
        0.5, 4000,
        "Minimal staging instance",
    )
    item(
        "Phase 1", "Infrastructure", "Azure Database for PostgreSQL Staging",
        "Managed database for development and testing",
        0.5, 2500,
        "Burstable B1ms tier",
    )
    item(
        "Phase 1", "Testing", "WhatsApp API Sandbox",
        "Meta WhatsApp Business API test numbers and sandbox",
        1, 5000,
        "Test message costs",
    )
    item(
        "Phase 1", "Testing", "API Integration Testing",
        "Gemini API test calls, SendGrid test emails, staging instance",
        1, 4000,
        "All third-party API test costs",
    )
    item(
        "Phase 1", "Testing", "Gemini Audio API Testing",
        "Voice note transcription and classification testing via Gemini Audio API",
        1, 1000,
        "Audio API test costs for voice complaints",
    )
    phase_subtotal("Phase 1")

    # ────────────────────────────────────────────────────────────
    # PHASE 2 — Pilot Deployment (Weeks 3-4)
    # ────────────────────────────────────────────────────────────
    item(
        "Phase 2", "Team", "Student Developer Stipends",
        "5 members during pilot: monitoring, bug fixes, officer onboarding",
        "=5*0.5", 25000,
        "Half-month stipend weeks 3-4",
    )
    item(
        "Phase 2", "Infrastructure", "Azure Container Apps",
        "Live backend for 5-ward pilot",
        0.5, 4000,
        "Small instance",
    )
    item(
        "Phase 2", "Infrastructure", "Azure Database for PostgreSQL",
        "Live database for pilot",
        0.5, 2500,
        "Burstable tier",
    )
    item(
        "Phase 2", "Infrastructure", "Azure Blob Storage",
        "Photo evidence for pilot complaints",
        0.5, 500,
        "Minimal at 1K complaints",
    )
    item(
        "Phase 2", "AI and ML", "Gemini 2.5 Flash-Lite API",
        "AI classification for 1,000 pilot complaints",
        1000, 0.05,
        "$0.10/M input + $0.40/M output @ 84 INR/USD, "
        "30% voice mix = Rs. 0.05/call",
    )
    item(
        "Phase 2", "Communication", "WhatsApp Business API",
        "Live messaging for 5-ward pilot",
        1000, 0.10,
        "Service convos FREE (citizen-initiated 24h window). "
        "Rs. 0.145/utility template × 2 notifs × 20% outside "
        "window = Rs. 0.058/complaint",
    )
    item(
        "Phase 2", "Communication", "SendGrid",
        "Department notification emails",
        1, 100,
        "Free tier covers 1K emails",
    )
    phase_subtotal("Phase 2")

    # ────────────────────────────────────────────────────────────
    # PHASE 3 — Partial Rollout (Weeks 5-6)
    # ────────────────────────────────────────────────────────────
    item(
        "Phase 3", "Team", "Student Developer Stipends",
        "Scaling support, QA, officer training, data ops",
        "=5*0.5", 25000,
        "Half-month stipend weeks 5-6",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Container Apps",
        "Scaled backend for 50 wards",
        0.5, 8000,
        "Medium instance",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Database for PostgreSQL Flexible",
        "Production-grade database with backups",
        0.5, 5000,
        "General Purpose tier",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Blob Storage",
        "40GB photo storage at 20K complaints",
        0.5, 1500,
        "LRS redundancy",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Monitor",
        "Observability and alerting",
        0.5, 1500,
        "Required for SLA",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Service Bus",
        "Webhook event queuing at 20K/month",
        0.5, 1000,
        "Standard tier",
    )
    item(
        "Phase 3", "AI and ML", "Gemini 2.5 Flash-Lite API",
        "AI classification for 20,000 complaints",
        20000, 0.015,
        "Rs. 0.012/call avg with voice mix, rounded up",
    )
    item(
        "Phase 3", "AI and ML", "Azure Machine Learning",
        "First retraining run on real pilot data",
        1, 8000,
        "Gradient Boosting retraining",
    )
    item(
        "Phase 3", "Communication", "WhatsApp Business API",
        "Messaging for 50-ward citizens",
        20000, 0.06,
        "2 utility templates × 20% outside 24h window "
        "× Rs. 0.145 = Rs. 0.058/complaint",
    )
    item(
        "Phase 3", "Communication", "SendGrid Essentials 50K",
        "$19.95/mo = Rs. 1,676/mo",
        0.5, 1676,
        "Essentials plan",
    )
    item(
        "Phase 3", "Infrastructure", "Azure Front Door CDN",
        "Officer dashboard CDN, domain, SSL",
        0.5, 1500,
        "Dashboard goes live",
    )
    phase_subtotal("Phase 3")

    # ────────────────────────────────────────────────────────────
    # PHASE 4 — Full Delhi Deployment (Weeks 7-8)
    # ────────────────────────────────────────────────────────────
    item(
        "Phase 4", "Team", "Student Developer Stipends",
        "Full-scale deployment support, ward onboarding, stress testing",
        "=5*0.5", 25000,
        "Half-month stipend weeks 7-8",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Container Apps",
        "Full-scale auto-scaling for 272 wards",
        0.5, 15000,
        "Production with zone redundancy",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Database for PostgreSQL",
        "High-availability production DB with geo-redundant backups "
        "and Indian data residency compliance",
        0.5, 12000,
        "Business Critical tier",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Blob Storage",
        "200GB photo storage at 1L complaints/month",
        0.5, 4000,
        "GRS geo-redundant",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Monitor",
        "Full production observability with custom dashboards",
        0.5, 3000,
        "90-day log retention",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Service Bus",
        "Premium tier guaranteed delivery at 1L/month",
        0.5, 2000,
        "Premium tier",
    )
    item(
        "Phase 4", "Infrastructure", "Azure Front Door CDN",
        "Full production CDN with WAF and DDoS protection",
        0.5, 3000,
        "WAF enabled",
    )
    item(
        "Phase 4", "AI and ML", "Gemini 2.5 Flash-Lite API",
        "Full-scale classification at 1,00,000 complaints/month",
        100000, 0.012,
        "Full scale with 30% voice mix",
    )
    item(
        "Phase 4", "AI and ML", "Azure Machine Learning",
        "Second retraining on expanded real complaint dataset",
        1, 8000,
        "Accuracy improves with real data",
    )
    item(
        "Phase 4", "Communication", "WhatsApp Business API",
        "Full Delhi citizen messaging at 1L conversations/month",
        100000, 0.06,
        "Peak communication cost",
    )
    item(
        "Phase 4", "Communication", "SendGrid Essentials 100K",
        "$34.95/mo = Rs. 2,940/mo",
        0.5, 2940,
        "Pro plan",
    )
    phase_subtotal("Phase 4")

    # ────────────────────────────────────────────────────────────
    # PHASE 5 — Steady State (monthly recurring)
    # ────────────────────────────────────────────────────────────
    item("Phase 5", "Infrastructure", "Azure Container Apps",
         "per month", 1, 15000, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Database for PostgreSQL",
         "per month", 1, 12000, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Blob Storage",
         "per month", 1, 4000, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Monitor",
         "per month", 1, 3000, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Service Bus",
         "per month", 1, 2000, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Front Door CDN",
         "per month", 1, 3000, "Recurring")
    item("Phase 5", "AI and ML", "Gemini 2.5 Flash-Lite API",
         "per month", 1, 1200, "Recurring")
    item("Phase 5", "AI and ML", "Azure Machine Learning",
         "per month amortized", 1, 5000, "Quarterly retraining")
    item("Phase 5", "Communication", "WhatsApp Business API",
         "per month", 1, 6000, "Recurring")
    item("Phase 5", "Communication", "SendGrid Essentials 100K",
         "per month", 1, 2940, "Recurring")
    item("Phase 5", "Infrastructure", "Azure backup and compliance",
         "per month", 1, 1500, "Recurring")
    item("Phase 5", "Infrastructure", "Azure Monitor log analytics",
         "per month", 1, 750, "Recurring")
    phase_subtotal("Phase 5")

    detail_end = row - 1

    # ── Borders on all detail rows ──
    for r in range(detail_start, detail_end + 1):
        _border_row(ws, r, 1, 8)

    # ── Zebra fill on item rows only (skip subtotals) ──
    alt = False
    for r in range(detail_start, detail_end + 1):
        if r in item_row_nums:
            if alt:
                _fill_row(ws, r, 1, 8, FILL_ALT)
            alt = not alt
        # subtotal rows already have FILL_SUBTOTAL set

    _thick_box(ws, s3, detail_end, 1, 8)

    # ================================================================
    # BACK-FILL Section 2 — Phase Cost formulas referencing Section 3
    # ================================================================
    for phase in ["Phase 1", "Phase 2", "Phase 3", "Phase 4", "Phase 5"]:
        cell = ws.cell(row=sum_rows[phase], column=6)
        _cur(cell, f"=G{sub_rows[phase]}")

    p1 = sub_rows["Phase 1"]
    p2 = sub_rows["Phase 2"]
    p3 = sub_rows["Phase 3"]
    p4 = sub_rows["Phase 4"]
    p5 = sub_rows["Phase 5"]

    # Grand Total (Phases 1-4) — formula
    _cur(ws.cell(row=gt_row, column=6), f"=G{p1}+G{p2}+G{p3}+G{p4}")
    ws.cell(row=gt_row, column=6).font = FONT_BOLD

    # Annual Maintenance — formula
    _cur(ws.cell(row=am_row, column=6), f"=G{p5}*12")
    ws.cell(row=am_row, column=6).font = FONT_BOLD

    row = detail_end + 3

    # ================================================================
    # SECTION 4 — UNIT ECONOMICS (all formulas)
    # ================================================================
    s4 = row
    _section_hdr(ws, row, 1, 8, "SECTION 4 — UNIT ECONOMICS")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(
        row=row, column=1,
        value="Unit Economics and Value Justification",
    ).font = FONT_BOLD_NAVY_12
    ws.cell(row=row, column=1).alignment = AL
    _border_row(ws, row, 1, 8)
    row += 1

    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    _table_hdr(ws, row, 1, 8)
    row += 1

    u_start = row

    def urow(metric: str, value, fmt: str = CURRENCY_FMT) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=metric)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        c = ws.cell(row=row, column=2, value=value)
        if isinstance(value, str) and not value.startswith("="):
            c.alignment = AL
        else:
            c.number_format = fmt
            c.alignment = AR
        _border_row(ws, row, 1, 8)
        row += 1

    urow("Total deployment cost Phases 1-4", f"=G{p1}+G{p2}+G{p3}+G{p4}")
    urow("Annual steady state cost Year 2", f"=G{p5}*12")
    urow("Cost per ward per month at full scale", f"=G{p5}/272", "0.00")
    urow("Cost per complaint at full scale", f"=G{p5}/100000", "0.00")
    urow("Citizens served estimate", 200000000, "#,##,##0")
    urow("Cost per citizen per year", f"=G{p5}*12/200000000", "0.00")
    urow("Wards covered at full deployment", 272, "#,##,##0")
    urow("Languages supported", "Hindi and English")
    urow("Complaints processed per year estimate", 1200000, "#,##,##0")

    u_end = row - 1

    # Zebra for unit economics
    for r in range(u_start, u_end + 1):
        if (r - u_start) % 2 == 1:
            _fill_row(ws, r, 1, 8, FILL_ALT)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
    ws.cell(
        row=row,
        column=1,
        value=(
            "Cost per complaint includes AI classification (with voice note "
            "support via Gemini 2.5 Flash-Lite), WhatsApp messaging, cloud "
            "infrastructure, email notifications, and ML-based "
            "auto-escalation — delivered at a fraction of the cost of "
            "traditional call-centre based grievance systems."
        ),
    ).alignment = AL
    for rr in (row, row + 1):
        _border_row(ws, rr, 1, 8)
    s4e = row + 1
    _thick_box(ws, s4, s4e, 1, 8)

    # ── Final formatting ──
    _col_widths(ws)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18

    return wb


# ─── Entry point ───────────────────────────────────────────────────

def main() -> Path:
    wb = build_workbook()
    out_path = Path(__file__).resolve().parent / OUTPUT_FILENAME
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        # Common on Windows if the XLSX is open in Excel.
        fallback = out_path.with_name(out_path.stem + "_NEW" + out_path.suffix)
        wb.save(fallback)
        return fallback


if __name__ == "__main__":
    p = main()
    print(f"Created: {p}")
